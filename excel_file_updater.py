import openpyxl


class ExcelFileUpdater:
    def __init__(self,
                 source_file,
                 source_sheet,
                 source_value_column_index,
                 source_content_column_index,
                 target_file,
                 target_sheet,
                 target_value_column_index,
                 target_content_column_index,
                 ):
        self.source_file = source_file
        self.source_sheet = source_sheet
        self.source_value_column_index = source_value_column_index
        self.source_content_column_index = source_content_column_index
        self.target_file = target_file
        self.target_sheet = target_sheet
        self.target_value_column_index = target_value_column_index
        self.target_content_column_index = target_content_column_index

    def update_target_file(self):

        source_workbook = openpyxl.load_workbook(self.source_file)
        source_sheet = source_workbook[self.source_sheet]

        target_workbook = openpyxl.load_workbook(self.target_file)
        target_sheet = target_workbook[self.target_sheet]

        for row in source_sheet.iter_rows(values_only=True):
            source_value = str(row[self.source_value_column_index])
            source_content = row[self.source_content_column_index]

            for target_row in target_sheet.iter_rows():
                target_value = str(target_row[self.target_value_column_index].value)

                if source_value == target_value:
                    target_row[self.target_content_column_index].value = source_content
                    break

        target_workbook.save(self.target_file)
        print("Update complete.")
