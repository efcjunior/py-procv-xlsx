import openpyxl

_source_file = 'files/source_file.xlsx'
_source_sheet = 'Sheet1'
_source_value_column_index = 0  # column 'A'
_source_content_column_index = 1  # column 'B'

_target_file = 'files/target_file.xlsx'
_target_sheet = 'BillingOffer'
_target_value_column_index = 8  # column 'I'
_target_content_column_index = 33  # column 'AH'


def update_target_file(
        param_source_file,
        param_source_sheet,
        param_source_value_column_index,
        param_source_content_column_index,
        param_target_file,
        param_target_sheet,
        param_target_value_column_index,
        param_target_content_column_index):
    # Load the source file
    source_workbook = openpyxl.load_workbook(param_source_file)
    source_sheet = source_workbook[param_source_sheet]

    # Load the target file
    target_workbook = openpyxl.load_workbook(param_target_file)
    target_sheet = target_workbook[param_target_sheet]

    # Iterate over the rows in the source sheet
    for row in source_sheet.iter_rows(values_only=True):
        source_value = str(row[param_source_value_column_index])  # Value in column 'A' of source file
        source_content = row[param_source_content_column_index]  # Value in column 'B' of source file

        # Iterate over the rows in the target sheet
        for target_row in target_sheet.iter_rows():
            target_value = str(target_row[param_target_value_column_index].value)  # Value in column 'I' of target file

            # If there's a match, update the corresponding cell in column 'AH' of target file
            if source_value == target_value:
                target_row[param_target_content_column_index].value = source_content
                break  # No need to continue searching

    # Save the updated target file
    target_workbook.save(param_target_file)
    print("Update complete.")


update_target_file(
        _source_file,
        _source_sheet,
        _source_value_column_index,
        _source_content_column_index,
        _target_file,
        _target_sheet,
        _target_value_column_index,
        _target_content_column_index)
